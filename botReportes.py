from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import smtplib
import schedule
import time
import sqlite3
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# Importa configuración desde archivo config.py
import config

# Cargar datos desde Excel o Base de Datos
def cargar_datos():
    if config.FUENTE_DATOS == 'excel':
        df = pd.read_excel(config.ARCHIVO_EXCEL)
    elif config.FUENTE_DATOS == 'bd':
        conn = sqlite3.connect(config.DB_PATH)
        df = pd.read_sql_query(config.DB_QUERY, conn)
        conn.close()
    else:
        raise ValueError("Fuente de datos no válida. Usa 'excel' o 'bd'.")
    return df

# Generar reporte PDF con resumen y tabla
def generar_reporte_pdf(nombre_pdf, df):
    doc = SimpleDocTemplate(nombre_pdf, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Título
    elements.append(Paragraph("Reporte de Ventas", styles['Title']))
    elements.append(Spacer(1, 12))

    # Resumen Ejecutivo
    total_ventas = df['cantidad'].sum()
    promedio_precio = df['precio_unitario'].mean()
    producto_mas_vendido = df.groupby('producto')['cantidad'].sum().idxmax()

    resumen_texto = (
        f"<b>Total unidades vendidas:</b> {total_ventas}<br/>"
        f"<b>Precio unitario promedio:</b> ${promedio_precio:.2f}<br/>"
        f"<b>Producto más vendido:</b> {producto_mas_vendido}<br/>"
    )
    elements.append(Paragraph("Resumen Ejecutivo", styles['Heading2']))
    elements.append(Paragraph(resumen_texto, styles['Normal']))
    elements.append(Spacer(1, 12))

    # Tabla agrupada por producto
    resumen = df.groupby('producto').agg({
        'cantidad': 'sum',
        'precio_unitario': 'mean'
    }).reset_index()

    table_data = [['Producto', 'Total Cantidad', 'Precio Promedio']] + resumen.values.tolist()
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))

    elements.append(Paragraph("Resumen por Producto", styles['Heading2']))
    elements.append(table)

    doc.build(elements)

# Generar Excel con hoja de resumen y hoja de datos con formato
def generar_reporte_excel(nombre_excel, df):
    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    total_ventas = df['cantidad'].sum()
    promedio_precio = df['precio_unitario'].mean()
    producto_mas_vendido = df.groupby('producto')['cantidad'].sum().idxmax()

    resumen_data = [
        ["Resumen Ejecutivo", ""],
        ["Total unidades vendidas", total_ventas],
        ["Precio unitario promedio", f"${promedio_precio:.2f}"],
        ["Producto más vendido", producto_mas_vendido],
    ]

    for row in resumen_data:
        ws_resumen.append(row)

    for cell in ws_resumen["A"] + ws_resumen[1]:
        cell.font = Font(bold=True)
    ws_resumen.column_dimensions["A"].width = 30
    ws_resumen.column_dimensions["B"].width = 30

    # Hoja de datos
    ws_datos = wb.create_sheet(title="Datos crudos")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_datos.append(r)

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws_datos[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for column_cells in ws_datos.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws_datos.column_dimensions[column_cells[0].column_letter].width = length + 2

    wb.save(nombre_excel)

# Generar ambos reportes y enviarlos
def generar_reportes():
    df = cargar_datos()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    nombre_excel = f'reporte_generado_{timestamp}.xlsx'
    nombre_pdf = f'reporte_generado_{timestamp}.pdf'

    generar_reporte_excel(nombre_excel, df)
    generar_reporte_pdf(nombre_pdf, df)

    print(f"[{timestamp}] Reportes generados: {nombre_excel}, {nombre_pdf}")
    enviar_correo([nombre_excel, nombre_pdf])

# Enviar correo con adjuntos
def enviar_correo(archivos):
    msg = MIMEMultipart()
    msg['From'] = config.REMITENTE
    msg['To'] = config.DESTINATARIO
    msg['Subject'] = 'Reportes automáticos generados por el bot'

    for archivo in archivos:
        with open(archivo, 'rb') as adj:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(adj.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{archivo}"')
            msg.attach(part)

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
        server.login(config.REMITENTE, config.CONTRASENA)
        server.sendmail(config.REMITENTE, config.DESTINATARIO, msg.as_string())

    print(f"Correo enviado a {config.DESTINATARIO} con los archivos adjuntos.")

# Iniciar el bot
def iniciar_bot():
    schedule.every(config.TIEMPO_REPORTE_MINUTOS).minutes.do(generar_reportes)
    print(f"Bot activo. Fuente: {config.FUENTE_DATOS}. Reportes cada {config.TIEMPO_REPORTE_MINUTOS} minutos...")
    while True:
        schedule.run_pending()
        time.sleep(1)

if __name__ == '__main__':
    iniciar_bot()
